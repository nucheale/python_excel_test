import openpyxl as xl
# from openpyxl import Workbook

# wb = Workbook()
# ws = wb.active



# wb = xl.load_workbook('C:/Users/Администратор/Desktop/Эко Плант ввоз.xlsx')
wb_name = 'Эко Плант ввоз.xlsx'
wb = xl.load_workbook(wb_name)
sh = wb['Гамма']
sh.cell(50, 10, 'Привет')
wb.save(wb_name)
# print(sh['A1'].value)
wb.close()

# # Data can be assigned directly to cells
# ws['A1'] = 42

# # Rows can also be appended
# ws.append([1, 2, 3])

# # Python types will automatically be converted
# import datetime
# ws['A2'] = datetime.datetime.now()

# # Save the file
# wb.save("sample.xlsx")